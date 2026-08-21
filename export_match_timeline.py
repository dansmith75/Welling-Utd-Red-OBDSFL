#!/usr/bin/env python3
"""Export the detailed Matchday event timeline used by the Dashboard Results info panel."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.time().isoformat() == "00:00:00":
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def normal(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def minute_number(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 9999.0


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python export_match_timeline.py /path/to/workbook.xlsx")

    workbook = Path(sys.argv[1]).expanduser().resolve()
    if not workbook.exists():
        raise FileNotFoundError(workbook)

    wb = load_workbook(workbook, data_only=True)
    if "MatchdayRecords" not in wb.sheetnames or "MatchdayRecords" not in wb["MatchdayRecords"].tables:
        output = []
    else:
        ws = wb["MatchdayRecords"]
        table = ws.tables["MatchdayRecords"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [ws.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        positions: dict[str, int] = {}
        for offset, header in enumerate(headers):
            key = normal(header)
            if key and key not in positions:
                positions[key] = min_col + offset

        def col(*names: str) -> Optional[int]:
            for name in names:
                found = positions.get(normal(name))
                if found is not None:
                    return found
            return None

        columns = {
            "sessionId": col("SessionId", "Session ID"),
            "matchId": col("MatchId", "Match ID"),
            "date": col("MatchDate", "Match Date"),
            "opposition": col("Opposition"),
            "competition": col("Competition"),
            "type": col("RecordType", "Record Type"),
            "playerId": col("PlayerId", "Player ID"),
            "player": col("DisplayName", "Display Name"),
            "relatedPlayerId": col("RelatedPlayerId", "Related Player ID"),
            "relatedPlayer": col("RelatedDisplayName", "Related Display Name"),
            "minute": col("Minute"),
            "detail": col("Detail"),
            "source": col("Source"),
        }

        grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
        ignored = {"", "session", "starter", "minutes", "assist"}

        for row_num in range(min_row + 1, max_row + 1):
            def value(name: str) -> Any:
                column = columns.get(name)
                return clean(ws.cell(row_num, column).value) if column is not None else None

            record_type = str(value("type") or "").strip()
            if record_type.lower() in ignored:
                continue

            date_value = value("date")
            opposition = value("opposition")
            match_id = value("matchId")
            if not date_value or not opposition:
                continue

            key = (str(match_id or ""), str(date_value), str(opposition))
            match = grouped.setdefault(key, {
                "matchId": match_id,
                "date": date_value,
                "opposition": opposition,
                "competition": value("competition"),
                "events": [],
            })

            match["events"].append({
                "type": record_type,
                "minute": value("minute"),
                "playerId": value("playerId"),
                "player": value("player"),
                "relatedPlayerId": value("relatedPlayerId"),
                "relatedPlayer": value("relatedPlayer"),
                "detail": value("detail"),
                "source": value("source"),
            })

        output = list(grouped.values())
        for match in output:
            match["events"].sort(key=lambda event: (minute_number(event.get("minute")), str(event.get("type") or "")))
        output.sort(key=lambda match: (str(match.get("date") or ""), str(match.get("opposition") or "")))

    destination = Path(__file__).resolve().parent / "data" / "timeline.json"
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {destination}")


if __name__ == "__main__":
    main()
