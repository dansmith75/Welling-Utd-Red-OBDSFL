r"""
Export Welling United Red OBDSFL workbook tables into clean dashboard JSON files.

Option A output: separate JSON files in ./data/
- players.json
- matches.json
- goals.json
- assists.json
- events.json
- attendance.json

How to run from the dashboard repo folder:
    python export_welling_json.py

Optional workbook override:
    python export_welling_json.py --workbook "C:\Users\dansm\OneDrive\Documents\Dan\Football\Welling United Red OBDSFL 26-27.xlsx"

Or set an environment variable:
    setx WELLING_WORKBOOK_PATH "C:\Users\dansm\OneDrive\Documents\Dan\Football\Welling United Red OBDSFL 26-27.xlsx"

Requirements:
    pip install openpyxl

Recommended setup:
- Keep the Excel workbook in OneDrive as the editable source of truth.
- Keep only the generated JSON files in the GitHub dashboard repo.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

TEAM = "Welling United Red OBDSFL"
SEASON = "2026/27"
WORKBOOK_NAME = "Welling United Red OBDSFL 26-27.xlsx"
DATA_DIR = Path("data")
ENV_WORKBOOK_PATH = "WELLING_WORKBOOK_PATH"


def candidate_workbook_paths(script_root: Path) -> List[Path]:
    """Possible workbook locations, in priority order."""
    candidates: List[Path] = []

    env_path = os.environ.get(ENV_WORKBOOK_PATH)
    if env_path:
        candidates.append(Path(env_path).expanduser())

    # Old/default behaviour: workbook in the dashboard repo root.
    candidates.append(script_root / WORKBOOK_NAME)

    # Recommended behaviour: workbook kept outside GitHub in OneDrive.
    candidates.append(Path.home() / "OneDrive" / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)
    candidates.append(Path.home() / "OneDrive - Personal" / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)
    candidates.append(Path.home() / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)

    # Remove duplicates while preserving order.
    unique: List[Path] = []
    seen = set()
    for path in candidates:
        resolved_key = str(path)
        if resolved_key not in seen:
            unique.append(path)
            seen.add(resolved_key)

    return unique


def resolve_workbook_path(script_root: Path, workbook_arg: Optional[str] = None) -> Path:
    """Find the workbook, either from --workbook, env var, or known default locations."""
    if workbook_arg:
        workbook_path = Path(workbook_arg).expanduser()
        if workbook_path.exists():
            return workbook_path
        raise FileNotFoundError(f"Workbook path from --workbook does not exist: {workbook_path}")

    candidates = candidate_workbook_paths(script_root)
    for path in candidates:
        if path.exists():
            return path

    candidate_text = "\n".join(f"- {path}" for path in candidates)
    raise FileNotFoundError(
        "Workbook not found. Checked these locations:\n"
        f"{candidate_text}\n\n"
        "Either move the workbook to one of those locations, run with --workbook, or set WELLING_WORKBOOK_PATH."
    )

# Columns to exclude from public player JSON. Keep contact details in Excel only.
PRIVATE_PLAYER_COLUMNS = {"number", "email"}


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def clean_value(value: Any) -> Any:
    """Convert Excel values into JSON-safe values."""
    if value is None:
        return None
    if isinstance(value, datetime):
        # Pure Excel dates often arrive as midnight datetimes.
        if value.time().isoformat() == "00:00:00":
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def camel_key(header: Any) -> str:
    """Turn Excel headers into predictable camelCase JSON keys."""
    text = str(header or "").strip()
    text = text.replace("/", " ").replace("-", " ")
    parts = re.findall(r"[A-Za-z0-9]+", text)
    if not parts:
        return ""
    first = parts[0].lower()
    rest = [p[:1].upper() + p[1:].lower() for p in parts[1:]]
    key = first + "".join(rest)
    replacements = {
        "id": "id",
        "displayname": "displayName",
        "playerid": "playerId",
        "sessionid": "sessionId",
        "sessionkey": "sessionKey",
        "sessiondate": "sessionDate",
        "sessiontype": "sessionType",
        "feepaid": "feePaid",
        "paymentstatus": "paymentStatus",
        "latepayment": "latePayment",
        "submittedby": "submittedBy",
        "submittedat": "submittedAt",
        "homeaway": "venue",
        "goalsfor": "goalsFor",
        "goalsagainst": "goalsAgainst",
    }
    return replacements.get(key, key)


def table_rows(workbook_path: Path, sheet_name: str, table_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Read an Excel table into a list of dictionaries."""
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    if table_name is None:
        if not ws.tables:
            return []
        table_name = next(iter(ws.tables.keys()))

    if table_name not in ws.tables:
        return []

    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [clean_value(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
    keys = [camel_key(header) for header in headers]

    rows: List[Dict[str, Any]] = []
    for row_num in range(min_row + 1, max_row + 1):
        row: Dict[str, Any] = {}
        has_data = False
        for col_num, key in zip(range(min_col, max_col + 1), keys):
            value = clean_value(ws.cell(row_num, col_num).value)
            row[key] = value
            if value not in (None, ""):
                has_data = True
        if has_data:
            rows.append(row)
    return rows


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {path}")


def export_players(workbook_path: Path) -> List[Dict[str, Any]]:
    rows = table_rows(workbook_path, "Squad", "Squad")
    players = []
    for row in rows:
        # Skip blank/total rows and keep contact details out of public JSON.
        player_id = row.get("id") or slugify(row.get("displayName") or row.get("name"))
        display_name = row.get("displayName") or row.get("name")
        if not player_id or not display_name:
            continue

        active_value = row.get("active")
        status_value = str(row.get("status") or "").strip().lower()
        active = bool(active_value) and status_value != "left"

        players.append({
            "id": player_id,
            "displayName": display_name,
            "active": active,
        })
    return players


def export_matches(workbook_path: Path) -> List[Dict[str, Any]]:
    rows = table_rows(workbook_path, "Fixtures", "Fixtures")
    matches = []
    for row in rows:
        if not row.get("date") and not row.get("opposition"):
            continue
        match_id = slugify(f"{row.get('date')}-{row.get('opposition')}")
        matches.append({
            "id": match_id,
            "date": row.get("date"),
            "day": row.get("day"),
            "opposition": row.get("opposition"),
            "competition": row.get("competition"),
            "venue": row.get("venue"),
            "postponed": bool(row.get("postponed")) if row.get("postponed") is not None else False,
            "goalsFor": row.get("goalsFor"),
            "goalsAgainst": row.get("goalsAgainst"),
            "result": row.get("result"),
        })
    return matches


def export_wide_player_stats(workbook_path: Path, sheet_name: str, output_key: str) -> List[Dict[str, Any]]:
    """Export wide match/player tables such as Goals, Assists and Events into row records."""
    rows = table_rows(workbook_path, sheet_name, sheet_name)
    output = []
    ignored = {"date", "opposition", "count"}
    for row in rows:
        date_value = row.get("date")
        opposition = row.get("opposition")
        if not date_value and not opposition:
            continue
        match_id = slugify(f"{date_value}-{opposition}")
        players: Dict[str, Any] = {}
        for key, value in row.items():
            if key in ignored:
                continue
            if value not in (None, "", 0):
                players[key] = value
        output.append({
            "matchId": match_id,
            "date": date_value,
            "opposition": opposition,
            output_key: players,
        })
    return output


def yes_no_to_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"yes", "true", "1", "paid"}:
        return True
    if text in {"no", "false", "0", "not paid"}:
        return False
    return None


def export_attendance(workbook_path: Path) -> Dict[str, Any]:
    rows = table_rows(workbook_path, "AttendanceRecords", "AttendanceRecords")

    sessions: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        session_key = row.get("sessionKey")
        player_id = row.get("playerId")
        display_name = row.get("displayName")
        status = row.get("status")
        if not session_key or not player_id or not status:
            continue

        if session_key not in sessions:
            sessions[session_key] = {
                "sessionKey": session_key,
                "sessionId": row.get("sessionId"),
                "date": row.get("sessionDate"),
                "type": row.get("sessionType"),
                "venue": row.get("venue"),
                "submittedBy": row.get("submittedBy"),
                "submittedAt": row.get("submittedAt"),
                "records": [],
            }

        record = {
            "recordKey": row.get("recordKey"),
            "playerId": player_id,
            "displayName": display_name,
            "status": status,
            "feePaid": yes_no_to_bool(row.get("feePaid")),
            "paymentStatus": row.get("paymentStatus"),
            "latePayment": yes_no_to_bool(row.get("latePayment")),
            "source": row.get("source"),
        }
        # Remove empty optional values to keep JSON tidy.
        record = {k: v for k, v in record.items() if v is not None}
        sessions[session_key]["records"].append(record)

    ordered_sessions = sorted(
        sessions.values(),
        key=lambda session: (session.get("date") or "", session.get("submittedAt") or ""),
    )

    return {
        "team": TEAM,
        "season": SEASON,
        "sessions": ordered_sessions,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Welling dashboard JSON from the Excel workbook.")
    parser.add_argument(
        "--workbook",
        help="Optional full path to the Excel workbook. If omitted, the script checks the repo root, WELLING_WORKBOOK_PATH, and common OneDrive locations.",
    )
    parser.add_argument(
        "--data-dir",
        default=str(DATA_DIR),
        help="Output data folder. Default: data",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    workbook_path = resolve_workbook_path(root, args.workbook)
    data_dir = root / Path(args.data_dir)

    print(f"Using workbook: {workbook_path}")
    print(f"Writing JSON to: {data_dir}")

    write_json(data_dir / "players.json", export_players(workbook_path))
    write_json(data_dir / "matches.json", export_matches(workbook_path))
    write_json(data_dir / "goals.json", export_wide_player_stats(workbook_path, "Goals", "goals"))
    write_json(data_dir / "assists.json", export_wide_player_stats(workbook_path, "Assists", "assists"))
    write_json(data_dir / "events.json", export_wide_player_stats(workbook_path, "Events", "events"))
    write_json(data_dir / "attendance.json", export_attendance(workbook_path))

    print("Done. Dashboard JSON files updated.")


if __name__ == "__main__":
    main()
